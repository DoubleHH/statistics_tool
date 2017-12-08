#coding:utf-8

# 将数据转换成excel

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl import Workbook
import csv
import codecs

def read_excel_sheet(excel_path, sheet_index):
    wb = load_workbook(excel_path)
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[sheet_index])
    return ws

def generate_excel(sheet_array, excel_name):
    wb = Workbook()
    sheet_index = 0
    for items_aray in sheet_array:
        if sheet_index == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = items_aray[0]
        generate_excel_sheet(ws, items_aray[1])
        sheet_index = sheet_index + 1
    wb.save(excel_name + '.xlsx')

def generate_csv(all_rows, csv_name):
    csv_file = open(csv_name + '.csv', 'wb')
    csv_file.write(codecs.BOM_UTF8)
    wr = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
    wr.writerows(all_rows)

def generate_excel_sheet(ws, items):
    row = 1
    columns = "ABCDEFGHIGKLMNOPQ"
    font = Font(size=14, color='FF000000')
    alignment = Alignment(horizontal='center', vertical='center')
    for item_array in items:
        column = 0
        ws.row_dimensions[row].height = 30
        for value in item_array:
            column_str = columns[column:(column+1)]
            ws.column_dimensions[column_str].width = 50
            cell_index = column_str + ("%d" % row)
            ws[cell_index] = value
            ws[cell_index].font = font
            ws[cell_index].alignment = alignment
            column = column + 1
        row = row + 1

def test_excel():
    titles = ["key值", "action动作", "note注释"]
    item1 = ["newtemplet.shopmenupg.catalog.content", "click", "KA模板分类点击事件"]
    item2 = ["newtemplet.shopmenupg.catalog.content", "click", "KA模板分类点击事件"]
    item3 = ["newtemplet.shopmenupg.catalog.content", "click", "KA模板分类点击事件"]
    items = [titles, item1, item2, item3]
    sheet_array = [
        ["same", items],
        ["iOS only", items],
        ["Android only", items]
    ]
    generate_excel(sheet_array, "test")
