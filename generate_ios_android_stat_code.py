#coding:utf-8

# 描述：把统计的excel表生成对应的 iOS 及 Android 代码

import os
import time
import glob
from sys import argv
import re
import operator
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import Alignment
import get_all_keys_from_code

ACTIONS = ['click', 'show', 'collect', 'ready']
COLORS  = {
    "blue": "0;34m",
    "green": "0;32m",
    "cyan": "0;36m",
    "red": "0;31m",
    "purple": "0;35m",
    "brown": "0;33m",
    "yellow": "1;33m",
    "lred": "1;31m",
}

def checkAction(action):
    global ACTIONS
    for value in ACTIONS:
        if cmp(value, action) == 0:
            return True
    return False

def print_error(index, string):
    print '\033[0;31m'
    error = ("Error! \nLine:%d: " % (index)) + string
    print error
    print '\033[0m'

def print_color_string(string, color):
    print ("\033[" + COLORS[color])
    print string
    print '\033[0m'

def read_excel_sheet(excel_path, sheet_index):
    wb = load_workbook(excel_path)
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[sheet_index])
    return ws

# 生成安卓统计代码
def generate_android_one_key_code(key, action, variable_name, note):
    dealed_key = key.replace("%", "%s")
    if cmp(action.lower(), "show") == 0:
        dealed_key = dealed_key + " " + "offline"
    code = "// " + note + "\n" + "String " + variable_name + " = \"" + dealed_key + "\";"
    return code

# 生成代码中的key值
def generate_variable_code(key, action):
    name = ("WM_STAT_" + key.upper().replace('.', '_') + "_" + action.upper())
    name = name.replace("%", "")
    return name

# 生成iOS统计.h代码
def generate_ios_one_key_code_h(variable_name, note):
    code = "/// " + note + "\n" + ("FOUNDATION_EXPORT NSString *const %s;" % (variable_name))
    return code

# 生成iOS统计.m代码
def generate_ios_one_key_code_m(key, action, variable_name, note):
    key = key.replace("%", "%@")
    offline = ""
    if cmp(action.lower(), "show") == 0:
        offline = " offline"
    code = "/// " + note + "\n" + ("NSString *const %s = @\"%s %s%s\";" % (variable_name, key, action.lower(), offline))
    return code

def string_with_value(value):
    if value == None:
        return ""
    else:
        return value.strip().strip('\n')

def generate_keys_array(ws, already_existed_keys):
    index = 0
    key_wappers = []
    android_code = ""
    ios_code_h = ""
    ios_code_m = ""
    # 当前excel生成的所有的代码key数组
    variable_array = []
    total_new_items = 0
    for row in ws.rows:
        index = index + 1
        row_array = []
        if index < 3:
            continue
        if index == 3:
            row_array.append("CODE_KEYS")
            for value in row:
                value_string = string_with_value(value.value)
                row_array.append(value_string)
            key_wappers.append(row_array)
            continue
        key = string_with_value(row[4].value)
        action = string_with_value(row[5].value)
        excel_note = string_with_value(row[3].value)
        excel_note = excel_note.replace("\n", "; ")
        note = string_with_value(row[1].value) + ", " + excel_note
        # print "key:%s, value:%s, note:%s" % (key, action, note)
        if len(key) == 0:
            print_error(index, "(%s) key为空" % (key))
            row_array.append("此行是成单统计")
        else:
            if checkAction(action) == False:
                print_error(index, "(%s) 统计action有问题，必须是click等" % (key))
            if len(note) == 0:
                print_error(index, "(%s) 注释为空" % (key))
            variable_name = generate_variable_code(key, action)
            append_string = ""
            if variable_array.count(variable_name) == 0:
                key_action = key + " " + action
                if already_existed_keys.count(key_action) != 0:
                    print_color_string(variable_name + " [ " + key_action + " ] SAME as in project", "lred")
                else:
                    variable_array.append(variable_name)
                    android_code = android_code + "\n" + generate_android_one_key_code(key, action, variable_name, note)
                    ios_code_h = ios_code_h + "\n" + generate_ios_one_key_code_h(variable_name, note)
                    ios_code_m = ios_code_m + "\n" + generate_ios_one_key_code_m(key, action, variable_name, note)
                    total_new_items = total_new_items + 1
            else:
                print_color_string(variable_name + " SAME in excel", "cyan")
            row_array.append(variable_name)


        for value in row:
            if value.value != None:
                row_array.append(value.value.strip().strip('\n'))
            else:
                row_array.append("")
        key_wappers.append(row_array)

    print "\n\nAndroid code:"
    print_color_string(android_code, "purple")

    print "\niOS code.h:"
    print_color_string(ios_code_h, "red")

    print "\niOS code.m:"
    print_color_string(ios_code_m, "brown")

    print_color_string("新的统计key共有%s个" % (total_new_items), "lred")
    return key_wappers

def generate_excel(sheet_array, excel_name):
    wb = Workbook()
    sheet_index = 0
    for items_aray in sheet_array:
        if sheet_index == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = items_aray[0]
        generate_excel_sheet(ws, items_aray[1])
        sheet_index = sheet_index + 1
    wb.save(excel_name + '.xlsx')

def generate_excel_sheet(ws, items):
    row = 1
    columns = "ABCDEFGHIGKLMNOPQ"
    font = Font(size=14, color='FF000000')
    alignment = Alignment(horizontal='center', vertical='center')
    for item_array in items:
        column = 0
        ws.row_dimensions[row].height = 30
        for value in item_array:
            column_str = columns[column:(column+1)]
            ws.column_dimensions[column_str].width = 50
            cell_index = column_str + ("%d" % row)
            ws[cell_index] = value
            ws[cell_index].font = font
            ws[cell_index].alignment = alignment
            column = column + 1
        row = row + 1

if __name__ == '__main__':
    if len(argv) < 2:
        print "Usage: python generate_keys.py excel_path ios_stat_file\n"
        exit(0)
    ws = read_excel_sheet(argv[1], 0)
    if ws == None:
        print_error(0, "sheet doesn't exist")
        exit(0)
    already_existed_keys = []
    if len(argv) > 2:
        already_existed_items = get_all_keys_from_code.generate_ios_keys(argv[2])
        already_existed_keys = already_existed_items.keys()
    array = generate_keys_array(ws, already_existed_keys)
    generate_excel([ ["keys", array], ], "4.4.0_stat_dealed")
